from time import gmtime, strftime, sleep
from datetime import datetime
import sqlite3
import sweepsupport as sws
import urllib
import sweepconfig

if sweepconfig.cte_export_ods:
    import pyoo
if sweepconfig.cte_export_openpyxl:
    from openpyxl import Workbook

if sweepconfig.cte_upload_web:
    import requests

# -*- coding: utf-8 -*-
"""
Automatic code generated by FoV Sweep Configurator tool
It creates a python script that commands the sweep actions and
triggers the frame acquistion shots for the configured FoV 
scanning action.
It is base (apart from python) in OpenCv.
To be successfully executed, you must provide local functions 
to give to this script the exact information needed (video source number)
and the functions to move the motors, this must included in a custom
module called "sweepsupport".
"""
if sweepconfig.cte_use_cvcam:
    import cv2

if sweepconfig.cte_use_photocam:
    import subprocess

if sweepconfig.cte_use_gphoto2:
    import gphoto2capture


def import_URL(URL):
    exec urllib.urlopen(URL).read() in globals()


# ###### Functions API ################


# ## Commands the motor to a given position
def commandMotor(x, y):
    if sweepconfig.cte_verbose:
       print ("Command Motor X: " + str(x) + " Y: " + str(y))
    return sws.commandMotor(x, y)


# ## Commands the motor to a given position
def commandMotorUnits(x, y, z):
    if sweepconfig.cte_verbose:
       print ("Command Motor in raw units X: " + str(x) + " Y: " + str(y) + " Z: " + str(z))
    return sws.commandMotorUnits(x, y, z)


# ## Investigate if the current movement has been executed
# ## you can also include here the user interaction, allowing
# ## him/her to quit the scanning operation
def stepDone():
    # Wait for command or movement time
    # returns are:
    #   -1 if the sweep operation must be cancelled
    #   1 if the movement has been done and the frame must be acquired
    #   0 does nothing, non blocking implementation is welcome
    return sws.stepDone()


sqlsentence = "INSERT INTO \"scan_ex_logs\" (\"step_order\", \"iteration\", \"step\", \"x\", \"y\", " + \
              "\"x_coord\", \"y_coord\", \"z_coord\", \"mx\", \"my\", \"mcomp\", \"mx_fdback\", \"my_fdback\", \"mcomp_fdback\", " + \
              "\"timestr\", \"scan_eng_run_id\", \"dtinit\", \"dtend\", \"created_at\", \"updated_at\") VALUES " + \
              "(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "

sqlprepare = "CREATE TABLE IF NOT EXISTS \"scan_ex_logs\" (\"id\" INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, " + \
             "\"mx\" float, \"my\" float, \"mcomp\" float, \"created_at\" datetime, \"updated_at\" datetime, " + \
             "\"scan_eng_run_id\" integer, \"step_order\" integer, \"iteration\" integer, \"step\" integer, \"x\" integer, \"y\" integer, " + \
             "\"x_coord\" float, \"y_coord\" float, \"z_coord\" float, \"timestr\" varchar, \"dtinit\" datetime, \"dtend\" datetime, " + \
             "\"mx_fdback\" float, \"my_fdback\" float, \"mcomp_fdback\" float);"

sqlsentence2 = "INSERT INTO \"scan_eng_runs\" (\"name\", \"scan_ex_id\", " + \
    "\"use_cam\", \"stab_time\", \"use_sim\", \"proto_rev\", " + \
    "\"ls1_va\", \"ls2_va\", \"ls3_va\", " + \
    "\"ls1_vh\", \"ls2_vh\", \"ls3_vh\", " + \
    "\"ls1_vi\", \"ls2_vi\", \"ls3_vi\", " + \
    "\"ls1_scale\", \"ls2_scale\", \"ls3_scale\", " + \
    "\"ls1_min\", \"ls2_min\", \"ls3_min\", " + \
    "\"ls1_max\", \"ls2_max\", \"ls3_max\", " + \
    "\"ls1_zero\", \"ls2_zero\", \"ls3_zero\", " + \
    "\"comp_factor_x\", \"comp_factor_y\", \"comp_divisor\", " + \
    "\"created_at\", \"updated_at\") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?," + \
    "?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "

sqlprepare2 = "CREATE TABLE IF NOT EXISTS \"scan_eng_runs\" (\"id\" INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, " + \
    "\"name\" varchar, \"created_at\" datetime, \"updated_at\" datetime, \"scan_ex_id\" integer, " + \
    "\"use_cam\" boolean, \"stab_time\" float, \"use_sim\" boolean, \"proto_rev\" integer, \"ls1_va\" float, " + \
    "\"ls2_va\" float, \"ls3_va\" float, \"ls1_vh\" float, \"ls2_vh\" float, \"ls3_vh\" float, \"ls1_vi\" float, " + \
    "\"ls2_vi\" float, \"ls3_vi\" float, \"ls1_scale\" float, \"ls2_scale\" float, \"ls3_scale\" float, " + \
    "\"ls1_min\" integer, \"ls2_min\" integer, \"ls3_min\" integer, \"ls1_max\" integer, \"ls2_max\" integer, " + \
    "\"ls3_max\" integer, \"ls1_zero\" integer, \"ls2_zero\" integer, \"ls3_zero\" integer, " + \
    "\"comp_factor_x\" float, \"comp_factor_y\" float, \"comp_divisor\" float);"
firstDbSentence = True
sweep_eng_run_id = None

engrunsheet = None
exlogsheet = None
doc = None
docrow = 0


def dbinsert(dbcon, cur_step_order, cur_iter, cur_step, step_x, step_y, step_x_coord, step_y_coord, step_z_coord, mx_setpoint, my_setpoint, mcomp_setpoint,
             mx_pos, my_pos, mcomp_pos, timestamp, dt_init, dt_end, ex_id, run_id):
    global firstDbSentence
    global docrow

    sdtcam = str(dt_end)
    sdtinit = str(dt_init)

    if firstDbSentence:
        item = [timestamp, ex_id,
                sweepconfig.cte_use_cvcam or sweepconfig.cte_use_photocam or sweepconfig.cte_use_gphoto2,
                sweepconfig.cte_stabilization_time, sweepconfig.cte_debug, sweepconfig.cte_proto_rev,
                sws.cte_vx, sws.cte_vy, sws.cte_vcomp, sws.cte_vhx, sws.cte_vhy, sws.cte_vhcomp,
                sws.cte_vix, sws.cte_viy, sws.cte_vicomp,
                sws.cte_lsx_scale, sws.cte_lsy_scale, sws.cte_lscomp_scale,
                sws.cte_lsx_min, sws.cte_lsy_min, sws.cte_lscomp_min,
                sws.cte_lsx_max, sws.cte_lsy_max, sws.cte_lscomp_max,
                sws.cte_lsx_zero, sws.cte_lsy_zero, sws.cte_lscomp_zero,
                sws.cte_comp_factor_x, sws.cte_comp_factor_y, sws.cte_comp_divisor,
                dt_end, dt_end]
        if sweepconfig.cte_verbose:
            print(sqlsentence2)
        dbcon.execute(sqlsentence2, item)
        run_id = dbcon.lastrowid
        firstDbSentence = False
        if sweepconfig.cte_export_ods:
            engrunsheet[0, 0:33].values = ["id", "name", "scan_ex_id", "use_cam", "stab_time",
                                          "use_sim", "proto_rev",
                                          "ls1_va", "ls2_va", "ls3_va",
                                          "ls1_vh", "ls2_vh", "ls3_vh",
                                          "ls1_vi", "ls2_vi", "ls3_vi",
                                          "ls1_scale", "ls2_scale", "ls3_scale",
                                          "ls1_min", "ls2_min", "ls3_min",
                                          "ls1_max", "ls2_max", "ls3_max",
                                          "ls1_zero", "ls2_zero", "ls3_zero",
                                          "comp_factor_x", "comp_factor_y", "comp_divisor",
                                          "created_at", "updated_at"]
            engrunsheet[1, 0].value = run_id
            item = [timestamp, ex_id,
                    str(sweepconfig.cte_use_cvcam or sweepconfig.cte_use_photocam or sweepconfig.cte_use_gphoto2),
                    sweepconfig.cte_stabilization_time, str(sweepconfig.cte_debug), sweepconfig.cte_proto_rev,
                    sws.cte_vx, sws.cte_vy, sws.cte_vcomp, sws.cte_vhx, sws.cte_vhy, sws.cte_vhcomp,
                    sws.cte_vix, sws.cte_viy, sws.cte_vicomp,
                    sws.cte_lsx_scale, sws.cte_lsy_scale, sws.cte_lscomp_scale,
                    sws.cte_lsx_min, sws.cte_lsy_min, sws.cte_lscomp_min,
                    sws.cte_lsx_max, sws.cte_lsy_max, sws.cte_lscomp_max,
                    sws.cte_lsx_zero, sws.cte_lsy_zero, sws.cte_lscomp_zero,
                    sws.cte_comp_factor_x, sws.cte_comp_factor_y, sws.cte_comp_divisor,
                    sdtcam, sdtcam]

            engrunsheet[1, 1:33].values = item
            exlogsheet[0, 0:21].values = ["id", "step_order", "iteration", "step", "x", "y", "x_coord", "y_coord", "z_coord", "mx", "my", "mcomp",
                                          "mx_fdback", "my_fdback", "mcomp_fdback", "timestr", "scan_eng_run_id",
                                          "dtinit", "dtend", "created_at", "updated_at"]
        if sweepconfig.cte_export_openpyxl:
            idx = 0
            item = ["id", "name", "scan_ex_id", "use_cam", "stab_time",
                                          "use_sim", "proto_rev",
                                          "ls1_va", "ls2_va", "ls3_va",
                                          "ls1_vh", "ls2_vh", "ls3_vh",
                                          "ls1_vi", "ls2_vi", "ls3_vi",
                                          "ls1_scale", "ls2_scale", "ls3_scale",
                                          "ls1_min", "ls2_min", "ls3_min",
                                          "ls1_max", "ls2_max", "ls3_max",
                                          "ls1_zero", "ls2_zero", "ls3_zero",
                                          "comp_factor_x", "comp_factor_y", "comp_divisor",
                                          "created_at", "updated_at"]
            for row in engrunsheet.iter_rows('A1:AG1'):
                for cell in row:
                    cell.value = item[idx]
                    idx += 1
            engrunsheet['A2'] = run_id
            idx = 0
            item = [timestamp, ex_id,
                    sweepconfig.cte_use_cvcam or sweepconfig.cte_use_photocam or sweepconfig.cte_use_gphoto2,
                    sweepconfig.cte_stabilization_time, sweepconfig.cte_debug, sweepconfig.cte_proto_rev,
                    sws.cte_vx, sws.cte_vy, sws.cte_vcomp, sws.cte_vhx, sws.cte_vhy, sws.cte_vhcomp,
                    sws.cte_vix, sws.cte_viy, sws.cte_vicomp,
                    sws.cte_lsx_scale, sws.cte_lsy_scale, sws.cte_lscomp_scale,
                    sws.cte_lsx_min, sws.cte_lsy_min, sws.cte_lscomp_min,
                    sws.cte_lsx_max, sws.cte_lsy_max, sws.cte_lscomp_max,
                    sws.cte_lsx_zero, sws.cte_lsy_zero, sws.cte_lscomp_zero,
                    sws.cte_comp_factor_x, sws.cte_comp_factor_y, sws.cte_comp_divisor,
                    sdtcam, sdtcam]
            for row in engrunsheet.iter_rows('B2:AG2'):
                for cell in row:
                    if sweepconfig.cte_verbose:
                        print("idx: " + str(idx))
                    cell.value = item[idx]
                    idx += 1

            item = ["id", "step_order", "iteration", "step", "x", "y", "x_coord", "y_coord", "z_coord", "mx", "my", "mcomp",
                    "mx_fdback", "my_fdback", "mcomp_fdback", "timestr", "scan_eng_run_id",
                    "dtinit", "dtend", "created_at", "updated_at"]
            idx = 0
            print "**************************************"
            for row in exlogsheet.iter_rows('A1:U1'):
                for cell in row:
                    cell.value = item[idx]
                    idx += 1

        # Web information upload
        if sweepconfig.cte_upload_web:
            postdata = {'sweep_eng_run': {"name": timestamp, "max_l1_speed": "33"}}
            r = requests.post(sweepconfig.cte_web_root + "/scan_eng_runs", data=postdata)
            print(r.status_code, r.reason)

        docrow += 1

    item = [cur_step_order, cur_iter, cur_step, step_x, step_y, step_x_coord, step_y_coord, step_z_coord, mx_setpoint, my_setpoint, mcomp_setpoint, mx_pos,
            my_pos, mcomp_pos, timestamp, run_id, dt_init, dt_end, dt_end, dt_end]
    dbcon.execute(sqlsentence, item)

    if sweepconfig.cte_export_ods:
        item = [cur_step_order, cur_iter, cur_step, step_x, step_y, step_x_coord, step_y_coord, step_z_coord, mx_setpoint, my_setpoint, mcomp_setpoint, mx_pos,
                my_pos, mcomp_pos, timestamp, run_id, sdtinit, sdtcam, sdtcam, sdtcam]
        exlogsheet[docrow, 0].value = dbcon.lastrowid
        exlogsheet[docrow, 1:21].values = item

    if sweepconfig.cte_export_openpyxl:
        item = [cur_step_order, cur_iter, cur_step, step_x, step_y, step_x_coord, step_y_coord, step_z_coord, mx_setpoint, my_setpoint, mcomp_setpoint, mx_pos,
                my_pos, mcomp_pos, timestamp, run_id, sdtinit, sdtcam, sdtcam, sdtcam]
        exlogsheet['A' + str(docrow + 1)] = dbcon.lastrowid
        idx = 0
        for row in exlogsheet.iter_rows('B' + str(docrow + 1) + ':U' + str(docrow + 1)):
            for cell in row:
                cell.value = item[idx]
                idx += 1

    docrow += 1

    return run_id


def dbprepare(dbcon):
    global engrunsheet
    global exlogsheet
    global doc

    dbcon.execute(sqlprepare2)
    dbcon.execute(sqlprepare)
    if sweepconfig.cte_export_ods:
        oodesktop = pyoo.Desktop('localhost', 2002)
        doc = oodesktop.create_spreadsheet()
        engrunsheet = doc.sheets.create('EngRun', index=1)
        exlogsheet = doc.sheets.create('ExLog', index=1)
        del doc.sheets[0]
    if sweepconfig.cte_export_openpyxl:
        doc = Workbook()
        engrunsheet = doc.create_sheet()
        engrunsheet.title = 'EngRun'
        exlogsheet = doc.create_sheet()
        exlogsheet.title = 'ExLog'
        ws = doc.active
        doc.remove_sheet(ws)
    return True


# ##### END Functions ############

# ##### Automatically generated code ###########

cte_z_y_exchange = False

cte_use_raw_units = False

sweep_ex_id = 15
steps = [ { 'i': (0),'c': (0),'x': (0),'y': (0),'x_coord': (0.0),'y_coord': (0.0),'z_coord': (0.0) },{ 'i': (0),'c': (1),'x': (0),'y': (1),'x_coord': (0.0110),'y_coord': (0.000),'z_coord': (0.0) }, ]

# ##### Automatically generated steps table ###########
# ##### END Automatically generated code ###########

# ### START EXECUTION ######

# Prepare the scan loop
curStep = 0
done = 0
# Create timestamp
timestr = strftime("%Y%m%d%H%M%S", gmtime())

if sweepconfig.cte_use_cvcam:
    # Cam has the video source
    cam = cv2.VideoCapture(sweepconfig.cte_camsource)
    if sweepconfig.cte_verbose:
        print ("Camera resolution:")
        print ("* Horizontal: " + str(cam.get(cv2.CAP_PROP_FRAME_WIDTH)))
        print ("* Vertical: " + str(cam.get(cv2.CAP_PROP_FRAME_HEIGHT)))

        # subprocess.check_call("exit 1", shell=True)

if sweepconfig.cte_disable_motors_first:
    sws.disableMotors()

if sweepconfig.cte_enable_motors_first:
    sws.enableMotors()

if sweepconfig.cte_reset_motors_first:
    sws.resetMotors()
    if sweepconfig.cte_verbose:
        print("Check motor positions after resets")
    sleep(sweepconfig.cte_stabilization_time)
    sws.motorPositions()
else:
    if sweepconfig.cte_verbose:
        print("Check initial motor positions")
    sleep(sweepconfig.cte_stabilization_time)
    sws.motorPositions()

# Prepare the Database
db = sqlite3.connect('./db/log.sqlite3')
if not db:
    ret = False
else:
    ret = True
    dbc = db.cursor()
    ret = dbprepare(dbc)

if not ret:
    done = -1
    print "Database ERROR! Aborting"

# Steps loop
# until ESC key is pressed
# or steps have finished
endStep = len(steps)

# endStep = 4
while (done != -1) and (curStep < endStep):
    # In stepX and stepY we have the step positions to be done
    iteration = steps[curStep]['i']
    step = steps[curStep]['c']
    stepX = steps[curStep]['x']
    stepY = steps[curStep]['y']
    stepXcoord = steps[curStep]['x_coord']
    stepYcoord = steps[curStep]['y_coord']
    stepZcoord = steps[curStep]['z_coord']
    # Command motor position for this step
    dtinit = datetime.now()
    if (not(cte_use_raw_units)):
        if (cte_z_y_exchange):
            done, mx, my, mcomp = commandMotor(stepXcoord, stepZcoord)
        else:
            done, mx, my, mcomp = commandMotor(stepXcoord, stepYcoord)
    else:
        if (cte_z_y_exchange):
            done, mx, my, mcomp = commandMotorUnits(stepXcoord, stepZcoord, stepYcoord)
        else:
            done, mx, my, mcomp = commandMotorUnits(stepXcoord, stepYcoord, stepZcoord)
    # Wait command to end
    while done == 0:
        done = stepDone()
    # END Command motor position for this step
    if done != -1:
        # Acquire image
        dtcam = datetime.now()
        # Wait some ms to stabilyze before reading position
        # not necessary if capture has been taken
        sleep(sweepconfig.cte_stabilization_time)
        capture_done = False
        if sweepconfig.cte_use_cvcam:
            ret, frame = cam.read()
            # save to disk
            strg = sweepconfig.cte_fileprefix + '%s_%03d_%03d_%03d.png' % (timestr, sweep_ex_id, iteration, step)
            cv2.imwrite(sweepconfig.cte_framePath + strg, frame)
            # show the image
            cv2.imshow('Current Frame', frame)
            capture_done = True
        if sweepconfig.cte_use_photocam:
            # We configure the image capture
            strg = 'D%s_%03d_%03d_%03d.jpg' % (timestr, sweep_ex_id, iteration, step)
            cmd = sweepconfig.cte_cameractrl_path + sweepconfig.cte_cameractrl_command
            args = sweepconfig.cte_cameractrl_filenamecmd + " " + strg
            if sweepconfig.cte_verbose:
                print("Photo set filename: " + cmd + " " + args)
            subprocess.check_output([cmd, args])
            args = sweepconfig.cte_cameractrl_capturecmd
            if sweepconfig.cte_verbose:
                print("Photo capture frame: " + cmd + " " + args)
            subprocess.check_output([cmd, args])
            capture_done = True
        if sweepconfig.cte_use_gphoto2:
            strg = sweepconfig.cte_gphoto2_filename_root + '%s_%03d_%03d_%03d.jpg' % (timestr, sweep_ex_id, iteration, step)
            gphoto2capture.capture(sweepconfig.cte_gphoto2_framePath, strg, False)
            capture_done = True

        # Shots a second picture after second stabilization time
        if sweepconfig.cte_second_picture:
            sleep(sweepconfig.cte_stabilization_time_pic2)
            if sweepconfig.cte_use_cvcam:
                ret, frame = cam.read()
                # save to disk
                strg = sweepconfig.cte_fileprefix + '%s_%03d_%03d_%03d_2.png' % (timestr, sweep_ex_id, iteration, step)
                cv2.imwrite(sweepconfig.cte_framePath + strg, frame)
                # show the image
                cv2.imshow('Current Frame', frame)
                capture_done = True
            if sweepconfig.cte_use_photocam:
                # We configure the image capture
                strg = 'D%s_%03d_%03d_%03d_2.jpg' % (timestr, sweep_ex_id, iteration, step)
                cmd = sweepconfig.cte_cameractrl_path + sweepconfig.cte_cameractrl_command
                args = sweepconfig.cte_cameractrl_filenamecmd + " " + strg
                if sweepconfig.cte_verbose:
                    print("Photo set filename: " + cmd + " " + args)
                subprocess.check_output([cmd, args])
                args = sweepconfig.cte_cameractrl_capturecmd
                if sweepconfig.cte_verbose:
                    print("Photo capture frame: " + cmd + " " + args)
                subprocess.check_output([cmd, args])
                capture_done = True
            if sweepconfig.cte_use_gphoto2:
                strg = sweepconfig.cte_gphoto2_filename_root + '%s_%03d_%03d_%03d_2.jpg' % (timestr, sweep_ex_id, iteration, step)
                gphoto2capture.capture(sweepconfig.cte_gphoto2_framePath, strg, False)
                capture_done = True

        # acquire the motor status
        mx_fdback, my_fdback, mcomp_fdback = sws.motorPositions()
        print ("Iteration: "+str(iteration)+"Step: "+str(step)+" Motor | mx: " + str(mx_fdback) + ", my: " + str(my_fdback) + ", mcomp: " + str(mcomp_fdback))
        # BD information store
        sweep_eng_run_id = dbinsert(dbc, curStep, iteration, step, stepX, stepY, stepXcoord, stepYcoord, stepZcoord, mx, my, mcomp, mx_fdback,
                                    my_fdback, mcomp_fdback, timestr, dtinit, dtcam, sweep_ex_id, sweep_eng_run_id)
        curStep += 1

# End of program, steps have finished or someone has cancelled the scan process
if curStep < len(steps) and sweepconfig.cte_verbose:
    # Scan process was cancelled
    print ("Scan process was cancelled")
    dummy = 0  # Dummy for avoiding indentation failures

db.commit()
db.close()

if sweepconfig.cte_export_ods:
    doc.save("./db/" + timestr + ".ods")
    doc.close()

if sweepconfig.cte_export_openpyxl:
    doc.save("./db/" + timestr + ".xlsx")

if sweepconfig.cte_use_cvcam:
    cam.release()
    cv2.destroyAllWindows()

sws.motorClose()
